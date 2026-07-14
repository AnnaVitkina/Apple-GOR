"""
Build the Conditions sheet with Port Code tables for the matrix workbook.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from build_matrix import DATA_START_ROW, SHIPMENT_COLUMNS, cell_text
from build_postal_code_zones import load_additional_info

CONDITIONS_SHEET_NAME = "conditions"
PORT_CODE_TABLE_NAME = "Port Code"
CONDITION_COLUMNS = ("Name", "Operator", "Values", "Scope")
PORT_CODE_OPERATOR = "equals to"
PORT_OF_LOADING_SCOPE = "Port of Loading"
PORT_OF_ENTRY_SCOPE = "Port of Entry"

PORT_OF_LOADING_COLUMN = "Port of Loading"
PORT_OF_ENTRY_COLUMN = "Port of Entry"

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
TABLE_TITLE_FILL = PatternFill("solid", fgColor="E2EFDA")
CONDITION_HIGHLIGHT_FILL = PatternFill("solid", fgColor="D9D9D9")
CONDITION_HIGHLIGHT_FONT = Font(underline="single")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

PORT_TOKEN_SPLIT_PATTERN = re.compile(r"[/\s]+")


@dataclass(frozen=True)
class PortCodeCondition:
    name: str
    operator: str
    values: str
    scope: str


def _style_header_cell(cell) -> None:
    cell.font = BOLD
    cell.fill = HEADER_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def _style_table_title_cell(cell) -> None:
    cell.font = BOLD
    cell.fill = TABLE_TITLE_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def _style_data_cell(cell) -> None:
    cell.font = NORMAL
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def extract_port_codes(additional_info_df: pd.DataFrame) -> list[tuple[str, str]]:
    """Return (Common Port(EN), Code) pairs from the Port Code table."""
    ports: list[tuple[str, str]] = []
    in_port_section = False
    header_seen = False

    for _, row in additional_info_df.iterrows():
        destination_city = cell_text(row.get("Destination City"))
        if destination_city == PORT_CODE_TABLE_NAME:
            in_port_section = True
            header_seen = False
            continue

        if not in_port_section:
            continue

        code = cell_text(row.get("Common Rating City"))
        national_code = cell_text(row.get("Destination City"))
        common_port_en = cell_text(row.get("Destination Country"))

        if not header_seen:
            if code == "Code" and common_port_en == "Common Port(EN)":
                header_seen = True
            continue

        if not code or not common_port_en or not national_code:
            continue
        if len(national_code) > 3 or len(code) > 8:
            continue

        ports.append((common_port_en, code))

    return ports


def _port_tokens(value: object) -> set[str]:
    text = cell_text(value)
    if not text:
        return set()

    tokens = {text.casefold()}
    for part in PORT_TOKEN_SPLIT_PATTERN.split(text):
        part_text = part.strip()
        if part_text:
            tokens.add(part_text.casefold())
    return tokens


def port_column_tokens(matrix_df: pd.DataFrame, column_name: str) -> set[str]:
    if column_name not in matrix_df.columns:
        return set()

    tokens: set[str] = set()
    for value in matrix_df[column_name]:
        tokens.update(_port_tokens(value))
    return tokens


def port_matches_column(
    common_port_en: str,
    code: str,
    column_tokens: set[str],
) -> bool:
    for candidate in (common_port_en, code):
        for token in _port_tokens(candidate):
            if token in column_tokens:
                return True
    return False


def build_port_code_conditions(
    additional_info_df: pd.DataFrame,
    matrix_df: pd.DataFrame,
) -> tuple[list[PortCodeCondition], list[PortCodeCondition]]:
    ports = extract_port_codes(additional_info_df)
    loading_tokens = port_column_tokens(matrix_df, PORT_OF_LOADING_COLUMN)
    entry_tokens = port_column_tokens(matrix_df, PORT_OF_ENTRY_COLUMN)

    loading_conditions: list[PortCodeCondition] = []
    entry_conditions: list[PortCodeCondition] = []
    seen_loading: set[str] = set()
    seen_entry: set[str] = set()

    for common_port_en, code in ports:
        if port_matches_column(common_port_en, code, loading_tokens):
            key = common_port_en.casefold()
            if key not in seen_loading:
                seen_loading.add(key)
                loading_conditions.append(
                    PortCodeCondition(
                        name=common_port_en,
                        operator=PORT_CODE_OPERATOR,
                        values=code,
                        scope=PORT_OF_LOADING_SCOPE,
                    )
                )

        if port_matches_column(common_port_en, code, entry_tokens):
            key = common_port_en.casefold()
            if key not in seen_entry:
                seen_entry.add(key)
                entry_conditions.append(
                    PortCodeCondition(
                        name=common_port_en,
                        operator=PORT_CODE_OPERATOR,
                        values=code,
                        scope=PORT_OF_ENTRY_SCOPE,
                    )
                )

    loading_conditions.sort(key=lambda item: item.name.casefold())
    entry_conditions.sort(key=lambda item: item.name.casefold())
    return loading_conditions, entry_conditions


def _write_port_code_table(
    worksheet,
    start_row: int,
    conditions: list[PortCodeCondition],
) -> int:
    title_cell = worksheet.cell(start_row, 1, PORT_CODE_TABLE_NAME)
    _style_table_title_cell(title_cell)

    header_row = start_row + 1
    for col_index, header in enumerate(CONDITION_COLUMNS, start=1):
        cell = worksheet.cell(header_row, col_index, header)
        _style_header_cell(cell)

    for offset, condition in enumerate(conditions, start=0):
        row_index = header_row + 1 + offset
        for col_index, value in enumerate(
            (condition.name, condition.operator, condition.values, condition.scope),
            start=1,
        ):
            cell = worksheet.cell(row_index, col_index, value)
            _style_data_cell(cell)

    return header_row + len(conditions)


def _condition_tokens(conditions: list[PortCodeCondition]) -> set[str]:
    tokens: set[str] = set()
    for condition in conditions:
        tokens.update(_port_tokens(condition.name))
        tokens.update(_port_tokens(condition.values))
    return tokens


def _cell_matches_conditions(cell_value: object, condition_tokens: set[str]) -> bool:
    if not condition_tokens:
        return False
    return bool(_port_tokens(cell_value) & condition_tokens)


def _port_column_index(column_name: str) -> int:
    return SHIPMENT_COLUMNS.index(column_name) + 1


def highlight_port_conditions_in_matrix(
    workbook: Workbook,
    loading_conditions: list[PortCodeCondition],
    entry_conditions: list[PortCodeCondition],
) -> int:
    worksheet = workbook["Rate card"]
    loading_col = _port_column_index(PORT_OF_LOADING_COLUMN)
    entry_col = _port_column_index(PORT_OF_ENTRY_COLUMN)
    loading_tokens = _condition_tokens(loading_conditions)
    entry_tokens = _condition_tokens(entry_conditions)
    highlighted = 0

    for row_index in range(DATA_START_ROW, worksheet.max_row + 1):
        loading_cell = worksheet.cell(row_index, loading_col)
        if _cell_matches_conditions(loading_cell.value, loading_tokens):
            loading_cell.fill = CONDITION_HIGHLIGHT_FILL
            loading_cell.font = CONDITION_HIGHLIGHT_FONT
            highlighted += 1

        entry_cell = worksheet.cell(row_index, entry_col)
        if _cell_matches_conditions(entry_cell.value, entry_tokens):
            entry_cell.fill = CONDITION_HIGHLIGHT_FILL
            entry_cell.font = CONDITION_HIGHLIGHT_FONT
            highlighted += 1

    return highlighted


def write_conditions_sheet(
    workbook: Workbook,
    matrix_df: pd.DataFrame,
    source_file: Path,
) -> tuple[int, int, int]:
    additional_info_df = load_additional_info(source_file)
    loading_conditions, entry_conditions = build_port_code_conditions(
        additional_info_df,
        matrix_df,
    )

    worksheet = workbook.create_sheet(title=CONDITIONS_SHEET_NAME)
    last_row = _write_port_code_table(worksheet, 1, loading_conditions)
    _write_port_code_table(worksheet, last_row + 2, entry_conditions)

    worksheet.column_dimensions["A"].width = 36.0
    worksheet.column_dimensions["B"].width = 14.0
    worksheet.column_dimensions["C"].width = 14.0
    worksheet.column_dimensions["D"].width = 18.0
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    highlighted = highlight_port_conditions_in_matrix(
        workbook,
        loading_conditions,
        entry_conditions,
    )

    return len(loading_conditions), len(entry_conditions), highlighted
