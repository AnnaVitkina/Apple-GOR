"""
Build Postal Code Zones txt from Additional Info common-rating tables.

Output columns (tab-separated):
  Name | Country | Postal Code | Excluded
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from build_matrix import (
    DATA_START_ROW,
    ORIGIN_CITY_LABEL_COLUMN,
    SHIPMENT_COLUMNS,
    cell_text,
)
from country_codes import to_country_iso
from project_paths import OUTPUT_DIR, PROCESSING_DIR, ensure_workspace_dirs
from us_ca_city_states import format_us_ca_postal_city

ORIGIN_COMMON_RATING_SECTION = "GOR25 Origin Common Rating"
DESTINATION_COMMON_RATING_SECTION = "GOR25 Destination Common Rating"

POSTAL_ZONE_COLUMNS = ("Name", "Country", "Postal Code", "Excluded")
POSTAL_CODE_ZONES_SUFFIX = "_postal_code_zones.txt"

CITY_HIGHLIGHT_FILL = PatternFill("solid", fgColor="DAEEF3")
CITY_HIGHLIGHT_FONT = Font(underline="single", bold=True)

DESTINATION_RAW_COLUMN = "Destination City"
DESTINATION_LABEL_COLUMN = "Destination City "
DESTINATION_LABEL_SUFFIX = " (Destination)"
ORIGIN_LABEL_SUFFIX = " (Origin)"


@dataclass(frozen=True)
class PostalCodeZone:
    name: str
    country: str
    postal_code: str
    excluded: str = "-"

    def as_txt_row(self) -> str:
        return "\t".join([self.name, self.country, self.postal_code, self.excluded])


def load_additional_info(file_path: Path) -> pd.DataFrame:
    workbook = pd.ExcelFile(file_path)
    if "Additional Info" not in workbook.sheet_names:
        raise RuntimeError(f"No 'Additional Info' sheet found in {file_path.name}")
    return pd.read_excel(file_path, sheet_name="Additional Info")


def _unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = cell_text(value)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def _normalize_country(value: object) -> str:
    text = cell_text(value)
    if not text:
        return ""
    iso = to_country_iso(text)
    if iso:
        return iso.upper()
    return text.upper()


def _build_zones_from_section(
    section_df: pd.DataFrame,
    *,
    city_column: str,
    country_column: str,
    suffix: str,
) -> list[PostalCodeZone]:
    if section_df.empty:
        return []

    grouped: dict[tuple[str, str], list[str]] = {}
    for _, row in section_df.iterrows():
        common_rating_city = cell_text(row.get("Common Rating City"))
        country = _normalize_country(row.get(country_column))
        city = cell_text(row.get(city_column))
        if not common_rating_city or not country or not city:
            continue

        key = (common_rating_city, country)
        grouped.setdefault(key, []).append(city)

    zones: list[PostalCodeZone] = []
    for (common_rating_city, country), cities in grouped.items():
        unique_cities = _unique_preserve_order(cities)
        postal_cities = [common_rating_city, *unique_cities]
        postal_cities = _unique_preserve_order(postal_cities)
        postal_values = [
            format_us_ca_postal_city(city, country, common_rating_city) for city in postal_cities
        ]
        zones.append(
            PostalCodeZone(
                name=f"{common_rating_city} ({suffix})",
                country=country,
                postal_code=", ".join(postal_values),
            )
        )

    return zones


def build_postal_code_zones(additional_info_df: pd.DataFrame) -> list[PostalCodeZone]:
    origin_df = additional_info_df[
        additional_info_df["Section Title"].astype(str).eq(ORIGIN_COMMON_RATING_SECTION)
    ]
    destination_df = additional_info_df[
        additional_info_df["Section Title"].astype(str).eq(DESTINATION_COMMON_RATING_SECTION)
    ]

    zones = [
        *_build_zones_from_section(
            origin_df,
            city_column="Origin City",
            country_column="Origin Country",
            suffix="Origin",
        ),
        *_build_zones_from_section(
            destination_df,
            city_column="Destination City",
            country_column="Destination Country",
            suffix="Destination",
        ),
    ]
    zones.sort(key=lambda zone: (zone.country, zone.name.casefold()))
    return zones


def _base_city_from_label(label: str, suffix: str) -> str:
    text = cell_text(label)
    if text.endswith(suffix):
        return text[: -len(suffix)].strip()
    return text


def _matrix_city_labels(
    matrix_df: pd.DataFrame,
    *,
    label_column: str,
    country_column: str,
) -> list[tuple[str, str]]:
    seen: set[str] = set()
    results: list[tuple[str, str]] = []

    for _, row in matrix_df.iterrows():
        if cell_text(row.get("Service Type")) == "SPECIAL":
            continue

        label = cell_text(row.get(label_column))
        country = _normalize_country(row.get(country_column))
        if not label or not country:
            continue

        key = label.casefold()
        if key in seen:
            continue
        seen.add(key)
        results.append((label, country))

    return results


def matrix_destination_labels(matrix_df: pd.DataFrame) -> list[tuple[str, str]]:
    """Return unique (Destination City label, country) pairs from the matrix."""
    return _matrix_city_labels(
        matrix_df,
        label_column=DESTINATION_LABEL_COLUMN,
        country_column="Destination Country",
    )


def matrix_origin_labels(matrix_df: pd.DataFrame) -> list[tuple[str, str]]:
    """Return unique (Origin City label, country) pairs from the matrix."""
    return _matrix_city_labels(
        matrix_df,
        label_column=ORIGIN_CITY_LABEL_COLUMN,
        country_column="Origin Country",
    )


def _append_matrix_city_zones(
    matrix_labels: list[tuple[str, str]],
    *,
    zones_by_name: dict[str, PostalCodeZone],
    label_suffix: str,
) -> list[PostalCodeZone]:
    city_zones: list[PostalCodeZone] = []
    for label, country in matrix_labels:
        if label in zones_by_name:
            city_zones.append(zones_by_name[label])
            continue

        base_city = _base_city_from_label(label, label_suffix)
        postal_code = (
            format_us_ca_postal_city(base_city, country, base_city)
            if country in {"US", "CA"}
            else base_city
        )
        city_zones.append(
            PostalCodeZone(
                name=label,
                country=country,
                postal_code=postal_code,
            )
        )
    return city_zones


def build_final_postal_zones(
    additional_info_df: pd.DataFrame,
    matrix_df: pd.DataFrame,
    *,
    include_origin_city: bool = False,
) -> list[PostalCodeZone]:
    all_zones = build_postal_code_zones(additional_info_df)
    zones_by_name = {zone.name: zone for zone in all_zones}

    final_zones: list[PostalCodeZone] = []
    if include_origin_city:
        final_zones.extend(
            _append_matrix_city_zones(
                matrix_origin_labels(matrix_df),
                zones_by_name=zones_by_name,
                label_suffix=ORIGIN_LABEL_SUFFIX,
            )
        )

    final_zones.extend(
        _append_matrix_city_zones(
            matrix_destination_labels(matrix_df),
            zones_by_name=zones_by_name,
            label_suffix=DESTINATION_LABEL_SUFFIX,
        )
    )

    final_zones.sort(key=lambda zone: (zone.country, zone.name.casefold()))
    return final_zones


def zone_names_for_suffix(zones: list[PostalCodeZone], suffix: str) -> set[str]:
    wrapped_suffix = f"({suffix})"
    return {zone.name for zone in zones if zone.name.endswith(wrapped_suffix)}


def destination_zone_names(zones: list[PostalCodeZone]) -> set[str]:
    return zone_names_for_suffix(zones, "Destination")


def origin_zone_names(zones: list[PostalCodeZone]) -> set[str]:
    return zone_names_for_suffix(zones, "Origin")


def write_postal_code_zones_txt(zones: list[PostalCodeZone], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["\t".join(POSTAL_ZONE_COLUMNS)]
    lines.extend(zone.as_txt_row() for zone in zones)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output_path


def _city_column_index(column_name: str) -> int:
    return SHIPMENT_COLUMNS.index(column_name) + 1


def highlight_labeled_city_column(
    worksheet,
    *,
    column_name: str,
    zone_names_set: set[str],
) -> int:
    if not zone_names_set or column_name not in SHIPMENT_COLUMNS:
        return 0

    col_index = _city_column_index(column_name)
    highlighted = 0

    for row_index in range(DATA_START_ROW, worksheet.max_row + 1):
        cell = worksheet.cell(row_index, col_index)
        value = cell_text(cell.value)
        if not value or value not in zone_names_set:
            continue

        cell.fill = CITY_HIGHLIGHT_FILL
        cell.font = CITY_HIGHLIGHT_FONT
        highlighted += 1

    return highlighted


def highlight_city_labels_in_matrix(
    matrix_path: Path,
    *,
    destination_zone_names_set: set[str],
    origin_zone_names_set: set[str] | None = None,
) -> tuple[int, int]:
    workbook = load_workbook(matrix_path)
    worksheet = workbook["Rate card"]

    destination_highlighted = highlight_labeled_city_column(
        worksheet,
        column_name=DESTINATION_LABEL_COLUMN,
        zone_names_set=destination_zone_names_set,
    )
    origin_highlighted = 0
    if origin_zone_names_set:
        origin_highlighted = highlight_labeled_city_column(
            worksheet,
            column_name=ORIGIN_CITY_LABEL_COLUMN,
            zone_names_set=origin_zone_names_set,
        )

    workbook.save(matrix_path)
    return destination_highlighted, origin_highlighted


def highlight_destination_cities_in_matrix(
    matrix_path: Path,
    destination_zone_names_set: set[str],
) -> int:
    destination_highlighted, _ = highlight_city_labels_in_matrix(
        matrix_path,
        destination_zone_names_set=destination_zone_names_set,
    )
    return destination_highlighted


def default_postal_zones_path(source_file: Path) -> Path:
    stem = source_file.stem.replace("_extracted", "")
    return OUTPUT_DIR / f"{stem}{POSTAL_CODE_ZONES_SUFFIX}"


def apply_postal_code_zones(
    *,
    source_file: Path,
    matrix_path: Path,
    matrix_df: pd.DataFrame,
    output_path: Path | None = None,
    include_origin_city: bool = False,
) -> tuple[Path, int, int, int, int]:
    additional_info_df = load_additional_info(source_file)
    all_zones = build_postal_code_zones(additional_info_df)
    final_zones = build_final_postal_zones(
        additional_info_df,
        matrix_df,
        include_origin_city=include_origin_city,
    )

    txt_path = output_path or default_postal_zones_path(source_file)
    write_postal_code_zones_txt(final_zones, txt_path)

    destination_names = destination_zone_names(final_zones)
    origin_names = origin_zone_names(final_zones) if include_origin_city else set()
    destination_highlighted, origin_highlighted = highlight_city_labels_in_matrix(
        matrix_path,
        destination_zone_names_set=destination_names,
        origin_zone_names_set=origin_names if include_origin_city else None,
    )
    return txt_path, len(final_zones), len(all_zones), destination_highlighted, origin_highlighted


def run_build_postal_code_zones(
    *,
    source_file: Path,
    matrix_path: Path,
    matrix_df: pd.DataFrame,
    output_path: Path | None = None,
    include_origin_city: bool = False,
) -> Path:
    txt_path, used_count, total_count, destination_highlighted, origin_highlighted = apply_postal_code_zones(
        source_file=source_file,
        matrix_path=matrix_path,
        matrix_df=matrix_df,
        output_path=output_path,
        include_origin_city=include_origin_city,
    )
    print(f"  Wrote {used_count} postal code zone(s) from {total_count} common-rating zone(s)")
    print(f"  Highlighted {destination_highlighted} destination label cell(s) in matrix")
    if include_origin_city:
        print(f"  Highlighted {origin_highlighted} origin label cell(s) in matrix")
    print(f"  Saved postal code zones to: {txt_path}")
    return txt_path


def list_extracted_files() -> list[Path]:
    if not PROCESSING_DIR.is_dir():
        return []
    return sorted(PROCESSING_DIR.glob("*_extracted.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
