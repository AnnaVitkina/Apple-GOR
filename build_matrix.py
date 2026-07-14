"""
Build a matrix workbook from extracted processing dataframes.

Matrix layout:
  - Shipment info (lane details)
  - Transport cost columns with multi-row headers
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from country_codes import to_country_iso
from project_paths import OUTPUT_DIR, PROCESSING_DIR, ensure_workspace_dirs

EXTRACTED_GLOB = "*_extracted.xlsx"

BASE_RATE_TAB = "base freight rates"
SERVICE_FEES_TAB = "service fees"
SPECIAL_SERVICE_TYPE = "SPECIAL"
APPLE_DISPOSITION_COLUMN = "Apple desposition"
APPROVED_DISPOSITION = "approved"

SERVICE_FEE_CHARGE_COLUMN = "Charge Code Description"
SERVICE_FEE_RATE_COLUMN = "Rate"
SERVICE_FEE_CURRENCY_COLUMN = "Curacncy"
SERVICE_FEE_ACCESSORIAL_COLUMN = "Accessorial Description"
SERVICE_FEE_ORIGINS_COLUMN = "Origins"
SERVICE_FEE_RATE_BASIS_COLUMN = "Rate Basis"

VALUE_COLUMN_PATTERN = re.compile(r"^Value\s*\(([^)]+)\)\s*$", re.IGNORECASE)
EXACT_CONTAINER_RATE_BASIS_PATTERN = re.compile(
    r"^\d+(?:RE|HCRE|HC|CN|CZ|HV)(?:/\d+(?:RE|HCRE|HC|CN|CZ|HV))*$",
    re.IGNORECASE,
)
RATE_BASIS_SEGMENT_PATTERN = re.compile(r"^(\d+)(RE|HCRE|HC|CN|CZ|HV)$", re.IGNORECASE)
EQUIPMENT_TYPE_ACCESSORIAL_PATTERN = re.compile(r"\d+'\s*(?:Std|HC)", re.IGNORECASE)

SPLIT_EQUIPMENT_TYPES: dict[str, tuple[str, ...]] = {
    "40' Std/HC Dry": ("40' Std Dry", "40' HC Dry"),
    "40' Std/HC Reefer": ("40' Std Reefer", "40' HC Reefer"),
}

EQUIPMENT_SORT_ORDER = (
    "20' Std Dry",
    "20' Std Reefer",
    "40' Std Dry",
    "40' Std Reefer",
    "40' HC Dry",
    "40' HC Reefer",
)

SHIPMENT_COLUMNS = (
    "Lane #",
    "Tab",
    "Service Type",
    "Origin Country",
    "Port of Loading",
    "Port of Loading ",
    "Destination Country",
    "Port of Entry",
    "Port of Entry ",
    "Destination City",
    "Destination City ",
    "Equipment Type",
)

COST_GROUP_ROW = 1
COST_NAME_ROW = 2
APPLY_IF_ROW = 3
RATE_BY_ROW = 4
COLUMN_HEADER_ROW = 5
DATA_START_ROW = 6

ACCESSORIAL_COSTS_SHEET_NAME = "Accessorial costs"
ACCESSORIAL_COSTS_COLUMNS = (
    "Name",
    "Price",
    "Currency",
    "Apply if",
    "Rate by",
)
ACCESSORIAL_HEADER_ROW = 1
ACCESSORIAL_DATA_START_ROW = 2
ACCESSORIAL_SERVICE_NOT_SPECIAL = "Service does not equal 'SPECIAL'"
BCL_TRUCKING_NAME_PATTERN = re.compile(
    r"Destination Trucking for BCL Multi Destinations \(([A-Z]{2})\)",
    re.IGNORECASE,
)

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
DUPLICATE_LANE_FILL = PatternFill("solid", fgColor="FFFF00")
TRANSPORT_GROUP_FILL = PatternFill("solid", fgColor="9BC2E6")
TRANSPORT_COST_FILL = PatternFill("solid", fgColor="BDD7EE")
SERVICE_FEE_COST_FILL = PatternFill("solid", fgColor="E2EFDA")
SERVICE_FEE_META_FILL = PatternFill("solid", fgColor="EAF4E3")
COST_META_FILL = PatternFill("solid", fgColor="F2F2F2")
SPECIAL_ROW_FILL = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RATE_NUMBER_FORMAT = "#,##0.00"
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


def cell_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def rate_value(value: object) -> float | int | None:
    if pd.isna(value):
        return None
    text = cell_text(value)
    if not text or text.lower() in {"on request", "n/a", "#n/a"}:
        return None
    try:
        number = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    if number.is_integer():
        return int(number)
    return number


def transport_cost_name(equipment_type: str) -> str:
    return f"Transport cost ({equipment_type})"


def expand_equipment_types(equipment: object) -> tuple[str, ...]:
    text = cell_text(equipment)
    if not text:
        return ()
    if text in SPLIT_EQUIPMENT_TYPES:
        return SPLIT_EQUIPMENT_TYPES[text]
    return (text,)


def container_rate_code(equipment_type: str) -> str:
    """Map equipment type to Container/{size}{suffix} (e.g. 40' HC Reefer -> 40HC)."""
    text = cell_text(equipment_type)
    size = "20" if text.startswith("20") else "40" if text.startswith("40") else ""
    lower = text.lower()

    if "hc reefer" in lower:
        suffix = "HC"
    elif "hc dry" in lower:
        suffix = "HV"
    elif "reefer" in lower:
        suffix = "CZ"
    elif "dry" in lower:
        suffix = "CN"
    else:
        suffix = ""

    return f"{size}{suffix}"


def rate_by_for_container_code(container_code: str) -> str:
    return f"Rate by: Container/{container_code}\nRegular rule"


def rate_basis_segment_to_container_code(segment: str) -> str:
    segment_text = cell_text(segment).upper()
    match = RATE_BASIS_SEGMENT_PATTERN.match(segment_text)
    if not match:
        return segment_text

    size, suffix_type = match.groups()
    if suffix_type == "RE":
        return f"{size}CZ"
    if suffix_type == "HCRE":
        return f"{size}HC"
    return f"{size}{suffix_type}"


def split_rate_basis_container_codes(rate_basis: str) -> list[str]:
    rate_basis_text = cell_text(rate_basis)
    if not EXACT_CONTAINER_RATE_BASIS_PATTERN.match(rate_basis_text):
        return []

    segments = [part.strip() for part in rate_basis_text.split("/") if part.strip()]
    if len(segments) <= 1:
        return []

    return [rate_basis_segment_to_container_code(segment) for segment in segments]


def rate_by_for_equipment(equipment_type: str) -> str:
    return f"Rate by: Container/{container_rate_code(equipment_type)}\nRegular rule"


def rate_by_for_cost_column(cost_column: str) -> str:
    prefix = "Transport cost ("
    suffix = ")"
    if cost_column.startswith(prefix) and cost_column.endswith(suffix):
        equipment_type = cost_column[len(prefix) : -len(suffix)]
        return rate_by_for_equipment(equipment_type)
    return "Rate by:\nRegular rule"


def is_transport_cost_column(column_name: str) -> bool:
    return column_name.startswith("Transport cost (")


def transport_cost_columns_from_df(matrix_df: pd.DataFrame) -> list[str]:
    return [column for column in matrix_df.columns if is_transport_cost_column(column)]


def service_fee_cost_columns_from_df(matrix_df: pd.DataFrame) -> list[str]:
    excluded = set(SHIPMENT_COLUMNS) | {"Currency"} | set(transport_cost_columns_from_df(matrix_df))
    return [column for column in matrix_df.columns if column not in excluded]


def all_cost_columns_from_df(matrix_df: pd.DataFrame) -> list[str]:
    return transport_cost_columns_from_df(matrix_df) + service_fee_cost_columns_from_df(matrix_df)


def find_value_column(df: pd.DataFrame) -> tuple[str, str] | None:
    for column in df.columns:
        match = VALUE_COLUMN_PATTERN.match(cell_text(column))
        if match:
            return str(column), match.group(1).strip().upper()
    return None


def sort_equipment_types(equipment_types: set[str]) -> list[str]:
    order_index = {name: index for index, name in enumerate(EQUIPMENT_SORT_ORDER)}

    def sort_key(name: str) -> tuple[int, str]:
        return (order_index.get(name, len(EQUIPMENT_SORT_ORDER)), name)

    return sorted(equipment_types, key=sort_key)


def collect_transport_cost_columns(rate_tabs: list[tuple[str, pd.DataFrame]]) -> list[str]:
    equipment_types: set[str] = set()

    for _, source_df in rate_tabs:
        if "Equipment Type" not in source_df.columns:
            continue
        for equipment in source_df["Equipment Type"].dropna().unique():
            equipment_types.update(expand_equipment_types(equipment))

    return [transport_cost_name(equipment) for equipment in sort_equipment_types(equipment_types)]


def list_extracted_files() -> list[Path]:
    return sorted(
        [
            path
            for path in PROCESSING_DIR.glob(EXTRACTED_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def prompt_selection(title: str, items: list[str]) -> int:
    print(f"\n{title}")
    for index, item in enumerate(items, start=1):
        print(f"  {index}. {item}")

    while True:
        raw = input("Enter number: ").strip()
        if not raw.isdigit():
            print("Please enter a valid number.")
            continue
        choice = int(raw)
        if 1 <= choice <= len(items):
            return choice - 1
        print("Number is out of range. Try again.")


def select_extracted_file(files: list[Path]) -> Path:
    if not files:
        print(f"No extracted files found in: {PROCESSING_DIR}")
        sys.exit(1)

    if len(files) == 1:
        print(f"\nUsing extracted file: {files[0].name}")
        return files[0]

    labels = [path.name for path in files]
    return files[prompt_selection("Select extracted file to build matrix from:", labels)]


def is_rate_tab(sheet_name: str) -> bool:
    lower = sheet_name.strip().lower()
    if lower == BASE_RATE_TAB:
        return True
    if "bcl" in lower:
        return True
    return False


def tab_equipment_type(sheet_name: str) -> str:
    lower = sheet_name.strip().lower()
    if lower == BASE_RATE_TAB:
        return "FCL"
    if "bcl" in lower:
        return "BCL"
    return ""


def destination_city_label(city: object) -> str:
    text = cell_text(city)
    if not text:
        return ""
    if text.lower().endswith("(destination)"):
        return text
    return f"{text} (Destination)"


def _column_or_blank(df: pd.DataFrame, column_name: str) -> pd.Series:
    if column_name not in df.columns:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="object")
    return df[column_name]


def load_service_fees_tab(file_path: Path) -> pd.DataFrame | None:
    workbook = pd.ExcelFile(file_path)
    for sheet_name in workbook.sheet_names:
        if sheet_name.strip().lower() == SERVICE_FEES_TAB:
            return pd.read_excel(file_path, sheet_name=sheet_name)
    return None


def _is_equipment_type_accessorial(accessorial: str) -> bool:
    return bool(EQUIPMENT_TYPE_ACCESSORIAL_PATTERN.search(cell_text(accessorial)))


def _is_per_exact_container_type_fee(row: pd.Series) -> bool:
    rate_basis = cell_text(row.get(SERVICE_FEE_RATE_BASIS_COLUMN))
    if EXACT_CONTAINER_RATE_BASIS_PATTERN.match(rate_basis):
        return True
    return _is_equipment_type_accessorial(row.get(SERVICE_FEE_ACCESSORIAL_COLUMN))


def _inherits_previous_charge(row: pd.Series) -> bool:
    if cell_text(row.get(SERVICE_FEE_CHARGE_COLUMN)):
        return False
    if cell_text(row.get(SERVICE_FEE_ORIGINS_COLUMN)):
        return True

    accessorial = cell_text(row.get(SERVICE_FEE_ACCESSORIAL_COLUMN))
    if re.search(r"NLRTM\s*-\s*\w+DC", accessorial, re.IGNORECASE):
        return True
    return _is_equipment_type_accessorial(accessorial)


def _trucking_country_from_accessorial(accessorial: str) -> str:
    compact = accessorial.upper().replace(" ", "")
    if "NLDC" in compact:
        return "NL"
    if "CZDC" in compact:
        return "CZ"
    if "ITDC" in compact:
        return "IT"
    if "UKDC" in compact:
        return "UK"
    return ""


def _destination_country_code(country_code: str) -> str:
    return "GB" if country_code.upper() == "UK" else country_code.upper()


def _trucking_country_from_name(name: str) -> str:
    match = BCL_TRUCKING_NAME_PATTERN.search(cell_text(name))
    return match.group(1).upper() if match else ""


def accessorial_apply_if(name: str) -> str:
    conditions: list[str] = []
    trucking_country = _trucking_country_from_name(name)
    if trucking_country:
        destination_country = _destination_country_code(trucking_country)
        conditions.append(
            f"Destination Country equals '{destination_country}' and Equipment Type contains 'BCL'"
        )
    conditions.append(ACCESSORIAL_SERVICE_NOT_SPECIAL)
    return " and ".join(conditions)


def accessorial_rate_by(charge: str, accessorial: str, rate_basis: str) -> str:
    accessorial_text = cell_text(accessorial)
    if _is_equipment_type_accessorial(accessorial_text):
        return rate_by_for_equipment(accessorial_text)

    rate_basis_text = cell_text(rate_basis)
    if EXACT_CONTAINER_RATE_BASIS_PATTERN.match(rate_basis_text):
        segments = [part.strip() for part in rate_basis_text.split("/") if part.strip()]
        if len(segments) == 1:
            container_code = rate_basis_segment_to_container_code(segments[0])
            return rate_by_for_container_code(container_code)

    if rate_basis_text:
        return f"Rate by: {rate_basis_text}\nRegular rule"
    return "Rate by:\nRegular rule"


def _collect_resolved_service_fee_items(
    service_fees_df: pd.DataFrame,
) -> tuple[list[dict[str, object]], dict[str, int]]:
    if service_fees_df.empty:
        return [], {}

    if APPLE_DISPOSITION_COLUMN not in service_fees_df.columns:
        return [], {}

    approved = service_fees_df[
        service_fees_df[APPLE_DISPOSITION_COLUMN]
        .astype(str)
        .str.strip()
        .str.lower()
        .eq(APPROVED_DISPOSITION)
    ].copy()

    resolved_rows: list[dict[str, object]] = []
    last_charge = ""
    for _, row in approved.iterrows():
        charge = cell_text(row.get(SERVICE_FEE_CHARGE_COLUMN))
        if charge:
            last_charge = charge
        elif _inherits_previous_charge(row):
            charge = last_charge
        accessorial = cell_text(row.get(SERVICE_FEE_ACCESSORIAL_COLUMN))
        origins = cell_text(row.get(SERVICE_FEE_ORIGINS_COLUMN))
        resolved_rows.append(
            {"charge": charge, "accessorial": accessorial, "origins": origins, "row": row}
        )

    charge_counts: dict[str, int] = {}
    for item in resolved_rows:
        charge = cell_text(item.get("charge"))
        if charge:
            charge_counts[charge] = charge_counts.get(charge, 0) + 1

    return resolved_rows, charge_counts


def service_fee_column_name(
    charge: str,
    accessorial: str,
    origins: str,
    *,
    charge_has_duplicates: bool,
) -> str:
    """Build service-fee cost column name from charge and accessorial descriptions."""
    charge = cell_text(charge)
    accessorial = cell_text(accessorial)
    origins = cell_text(origins)
    charge_lower = charge.lower()
    accessorial_lower = accessorial.lower()

    if not charge and accessorial_lower.startswith("gps"):
        return "GPS"

    if "waiting hours" in charge_lower or "driver waiting time" in charge_lower:
        if "origin" in accessorial_lower:
            return "Waiting Time (Origin)"
        if "destination" in accessorial_lower:
            return "Waiting Time (Destination)"
        return "Waiting Time"

    if charge_lower == "cancellation fee":
        if "prior to dispatch" in accessorial_lower and "picked up" in accessorial_lower:
            return "Cancellation Fee (prior to dispatch)"
        if "dispatched to oem" in accessorial_lower:
            return "Cancellation Fee (after dispatch)"
        if origins:
            return f"Cancellation Fee ({origins})"
        return charge

    if charge_lower == "customs fee":
        if "electronic entry fee" in accessorial_lower:
            return "Customs Entry (Australia - Customs Electronic Entry Fee)"
        if "clearance fee" in accessorial_lower:
            return "Customs Clearance Fee(Australia - Customs Clearance Fee)"
        return charge

    if "destination trucking for bcl" in charge_lower:
        country = _trucking_country_from_accessorial(accessorial)
        if country:
            return f"Destination Trucking for BCL Multi Destinations ({country})"
        return "Destination Trucking for BCL Multi Destinations"

    if "import/export handling fee" in charge_lower:
        if origins:
            return f"Handling (Import/Export Handling Fee ({origins}))"
        return "Handling (Import/Export Handling Fee)"

    if "hanoi origin consolid" in charge_lower and accessorial:
        return f"{charge} ({accessorial})"

    if "destination de-consolidation" in charge_lower and accessorial:
        return f"{charge} ({accessorial})"

    if charge:
        if charge_has_duplicates and accessorial and accessorial_lower != charge_lower:
            return f"{charge} ({accessorial})"
        return charge

    if accessorial:
        return accessorial

    if origins:
        return origins

    return ""


def _ensure_unique_column_name(name: str, used_names: set[str]) -> str:
    if name not in used_names:
        used_names.add(name)
        return name

    for suffix in range(2, 1000):
        candidate = f"{name} ({suffix})"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate

    raise RuntimeError(f"Could not create a unique service fee column name for: {name}")


def extract_approved_service_fees(service_fees_df: pd.DataFrame) -> list[tuple[str, float | int, str, str]]:
    """Return (column_name, rate, currency, rate_by) for approved service fees."""
    resolved_rows, charge_counts = _collect_resolved_service_fee_items(service_fees_df)
    if not resolved_rows:
        return []

    used_names: set[str] = set()
    fees: list[tuple[str, float | int, str, str]] = []

    for item in resolved_rows:
        row = item["row"]
        if _is_per_exact_container_type_fee(row):
            continue

        charge = cell_text(item.get("charge"))
        accessorial = cell_text(item.get("accessorial"))
        origins = cell_text(item.get("origins"))
        rate = rate_value(row.get(SERVICE_FEE_RATE_COLUMN))
        currency = cell_text(row.get(SERVICE_FEE_CURRENCY_COLUMN)).upper()

        column_name = service_fee_column_name(
            charge,
            accessorial,
            origins,
            charge_has_duplicates=charge_counts.get(charge, 0) > 1,
        )
        if not column_name or rate is None:
            continue

        column_name = _ensure_unique_column_name(column_name, used_names)
        rate_basis = cell_text(row.get(SERVICE_FEE_RATE_BASIS_COLUMN))
        rate_by = f"Rate by: {rate_basis}\nRegular rule" if rate_basis else "Rate by:\nRegular rule"
        fees.append((column_name, rate, currency, rate_by))

    return fees


def extract_accessorial_costs(
    service_fees_df: pd.DataFrame,
) -> list[tuple[str, float | int, str, str, str]]:
    """Return (name, price, currency, apply_if, rate_by) for the Accessorial costs tab."""
    resolved_rows, charge_counts = _collect_resolved_service_fee_items(service_fees_df)
    if not resolved_rows:
        return []

    used_names: set[str] = set()
    costs: list[tuple[str, float | int, str, str, str]] = []

    for item in resolved_rows:
        row = item["row"]
        charge = cell_text(item.get("charge"))
        accessorial = cell_text(item.get("accessorial"))
        origins = cell_text(item.get("origins"))
        rate = rate_value(row.get(SERVICE_FEE_RATE_COLUMN))
        currency = cell_text(row.get(SERVICE_FEE_CURRENCY_COLUMN)).upper()
        rate_basis = cell_text(row.get(SERVICE_FEE_RATE_BASIS_COLUMN))

        column_name = service_fee_column_name(
            charge,
            accessorial,
            origins,
            charge_has_duplicates=charge_counts.get(charge, 0) > 1,
        )
        if not column_name or rate is None:
            continue

        container_codes = split_rate_basis_container_codes(rate_basis)
        if len(container_codes) > 1:
            for container_code in container_codes:
                split_name = _ensure_unique_column_name(
                    f"{column_name} (Container/{container_code})",
                    used_names,
                )
                costs.append(
                    (
                        split_name,
                        rate,
                        currency,
                        accessorial_apply_if(split_name),
                        rate_by_for_container_code(container_code),
                    )
                )
            continue

        column_name = _ensure_unique_column_name(column_name, used_names)
        costs.append(
            (
                column_name,
                rate,
                currency,
                accessorial_apply_if(column_name),
                accessorial_rate_by(charge, accessorial, rate_basis),
            )
        )

    return costs


def build_special_lane_row(
    lane_number: int,
    transport_cost_columns: list[str],
    service_fees: list[tuple[str, float | int, str, str]],
) -> dict[str, object]:
    row: dict[str, object] = {column: "" for column in SHIPMENT_COLUMNS}
    row["Lane #"] = lane_number
    row["Service Type"] = SPECIAL_SERVICE_TYPE
    row["Currency"] = ""

    for column in transport_cost_columns:
        row[column] = None

    for column_name, rate, _currency, _rate_by in service_fees:
        row[column_name] = rate

    return row


def append_special_lane(
    matrix_df: pd.DataFrame,
    service_fees_df: pd.DataFrame | None,
) -> tuple[pd.DataFrame, dict[str, str], dict[str, str], list[tuple[str, float | int, str, str]]]:
    """Append SPECIAL lane and service-fee cost columns.

    Returns df, rate_by map, per-fee currency map, and extracted service fees.
    """
    if service_fees_df is None or service_fees_df.empty:
        return matrix_df, {}, {}, []

    service_fees = extract_approved_service_fees(service_fees_df)
    if not service_fees:
        return matrix_df, {}, {}, []

    transport_cost_columns = transport_cost_columns_from_df(matrix_df)
    service_fee_columns = [name for name, _, _, _ in service_fees]
    rate_by_map = {name: rate_by for name, _, _, rate_by in service_fees}
    currency_map = {name: currency for name, _, currency, _ in service_fees}

    result = matrix_df.copy()
    for column in service_fee_columns:
        if column not in result.columns:
            result[column] = None

    lane_number = int(result["Lane #"].max()) + 1 if not result.empty else 1
    special_row = build_special_lane_row(lane_number, transport_cost_columns, service_fees)
    result = pd.concat([result, pd.DataFrame([special_row])], ignore_index=True)

    ordered_columns = [
        *SHIPMENT_COLUMNS,
        "Currency",
        *transport_cost_columns,
        *service_fee_columns,
    ]
    return result.loc[:, ordered_columns], rate_by_map, currency_map, service_fees


def write_accessorial_costs_sheet(
    workbook: Workbook,
    accessorial_costs: list[tuple[str, float | int, str, str, str]],
) -> None:
    worksheet = workbook.create_sheet(title=ACCESSORIAL_COSTS_SHEET_NAME)

    for col_index, header in enumerate(ACCESSORIAL_COSTS_COLUMNS, start=1):
        cell = worksheet.cell(ACCESSORIAL_HEADER_ROW, col_index, header)
        _style_header_cell(cell, center=header in {"Price", "Currency"})

    for row_offset, (name, price, currency, apply_if, rate_by) in enumerate(
        accessorial_costs,
        start=0,
    ):
        excel_row = ACCESSORIAL_DATA_START_ROW + row_offset

        name_cell = worksheet.cell(excel_row, 1, name)
        name_cell.alignment = LEFT
        name_cell.border = THIN_BORDER

        price_cell = worksheet.cell(excel_row, 2, price)
        price_cell.number_format = RATE_NUMBER_FORMAT
        price_cell.alignment = CENTER
        price_cell.border = THIN_BORDER

        currency_cell = worksheet.cell(excel_row, 3, currency)
        currency_cell.alignment = CENTER
        currency_cell.border = THIN_BORDER

        apply_if_cell = worksheet.cell(excel_row, 4, apply_if)
        apply_if_cell.alignment = LEFT
        apply_if_cell.border = THIN_BORDER

        rate_by_cell = worksheet.cell(excel_row, 5, rate_by)
        rate_by_cell.alignment = LEFT
        rate_by_cell.border = THIN_BORDER

    worksheet.column_dimensions["A"].width = 52.0
    worksheet.column_dimensions["B"].width = 12.0
    worksheet.column_dimensions["C"].width = 12.0
    worksheet.column_dimensions["D"].width = 48.0
    worksheet.column_dimensions["E"].width = 34.0

    if accessorial_costs:
        for row_offset, (name, *_rest) in enumerate(accessorial_costs, start=0):
            excel_row = ACCESSORIAL_DATA_START_ROW + row_offset
            worksheet.row_dimensions[excel_row].height = 36.0
            width = _column_width_for_header(name)
            if width > worksheet.column_dimensions["A"].width:
                worksheet.column_dimensions["A"].width = min(width, 60.0)
            apply_if_width = _column_width_for_header(accessorial_costs[row_offset][3])
            if apply_if_width > worksheet.column_dimensions["D"].width:
                worksheet.column_dimensions["D"].width = min(apply_if_width, 72.0)

    worksheet.freeze_panes = worksheet.cell(ACCESSORIAL_DATA_START_ROW, 1)
    worksheet.sheet_view.showGridLines = False


def duplicate_lane_row_indices(matrix_df: pd.DataFrame) -> set[int]:
    """Matrix row indices for lanes that share shipment identity with another lane."""
    grouped: dict[tuple[str, ...], list[int]] = {}

    for index, row in matrix_df.iterrows():
        if cell_text(row.get("Service Type")) == SPECIAL_SERVICE_TYPE:
            continue
        key = _lane_identity_key(row)
        grouped.setdefault(key, []).append(int(index))

    duplicated: set[int] = set()
    for indices in grouped.values():
        if len(indices) > 1:
            duplicated.update(indices)
    return duplicated


def load_rate_tabs(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    workbook = pd.ExcelFile(file_path)
    loaded: list[tuple[str, pd.DataFrame]] = []

    for sheet_name in workbook.sheet_names:
        if not is_rate_tab(sheet_name):
            continue
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if df.empty:
            continue
        loaded.append((sheet_name, df))

    return loaded


def count_source_lane_rows(rate_tabs: list[tuple[str, pd.DataFrame]]) -> int:
    return sum(len(source_df) for _, source_df in rate_tabs)


def build_shipment_and_cost_rows(rate_tabs: list[tuple[str, pd.DataFrame]]) -> pd.DataFrame:
    transport_cost_columns = collect_transport_cost_columns(rate_tabs)
    rows: list[dict[str, object]] = []
    lane_number = 0

    for tab_name, source_df in rate_tabs:
        port_of_loading = _column_or_blank(source_df, "Port of Loading")
        port_of_entry = _column_or_blank(source_df, "Port of Entry")
        destination_city = _column_or_blank(source_df, "Destination City")
        value_column_info = find_value_column(source_df)

        for row_idx, source_row in source_df.iterrows():
            lane_number += 1
            equipment = source_row.get("Equipment Type")
            expanded_equipment = expand_equipment_types(equipment)

            rate_amount = None
            currency = ""
            if value_column_info is not None:
                value_column, currency_code = value_column_info
                rate_amount = rate_value(source_row.get(value_column))
                currency = currency_code if rate_amount is not None else ""

            matrix_row: dict[str, object] = {
                "Lane #": lane_number,
                "Tab": tab_name,
                "Service Type": cell_text(source_row.get("Service Type")),
                "Origin Country": cell_text(source_row.get("Origin Country")),
                "Port of Loading": cell_text(port_of_loading.loc[row_idx]),
                "Port of Loading ": cell_text(port_of_loading.loc[row_idx]),
                "Destination Country": to_country_iso(source_row.get("Destination Country")),
                "Port of Entry": cell_text(port_of_entry.loc[row_idx]),
                "Port of Entry ": cell_text(port_of_entry.loc[row_idx]),
                "Destination City": cell_text(destination_city.loc[row_idx]),
                "Destination City ": destination_city_label(destination_city.loc[row_idx]),
                "Equipment Type": tab_equipment_type(tab_name),
                "Currency": currency,
            }

            for cost_column in transport_cost_columns:
                matrix_row[cost_column] = None

            for equipment_type in expanded_equipment:
                matrix_row[transport_cost_name(equipment_type)] = rate_amount

            rows.append(matrix_row)

    if not rows:
        return pd.DataFrame(columns=[*SHIPMENT_COLUMNS, "Currency", *transport_cost_columns])

    matrix_df = pd.DataFrame(rows)
    return merge_unique_lanes(matrix_df)


LANE_IDENTITY_COLUMNS = (
    "Service Type",
    "Origin Country",
    "Port of Loading",
    "Port of Loading ",
    "Port of Entry",
    "Port of Entry ",
    "Destination City",
    "Equipment Type",
)


def _lane_identity_key(row: pd.Series) -> tuple[str, ...]:
    return tuple(cell_text(row.get(column)) for column in LANE_IDENTITY_COLUMNS)


def _non_null_cost_values(series: pd.Series) -> list[float | int]:
    values: list[float | int] = []
    for item in series:
        parsed = rate_value(item)
        if parsed is not None:
            values.append(parsed)
    return values


def _can_merge_rows(
    row_a: dict[str, object],
    row_b: dict[str, object],
    transport_cost_columns: list[str],
) -> bool:
    for column in transport_cost_columns:
        values = _non_null_cost_values(pd.Series([row_a.get(column), row_b.get(column)]))
        if len(values) > 1 and len(set(values)) > 1:
            return False
    return True


def _merge_two_rows(
    row_a: dict[str, object],
    row_b: dict[str, object],
    transport_cost_columns: list[str],
) -> dict[str, object]:
    merged = dict(row_a)

    for column in transport_cost_columns:
        values = _non_null_cost_values(pd.Series([row_a.get(column), row_b.get(column)]))
        merged[column] = values[0] if values else None

    currencies = [
        cell_text(value)
        for value in (row_a.get("Currency"), row_b.get("Currency"))
        if cell_text(value)
    ]
    unique_currencies = list(dict.fromkeys(currencies))
    merged["Currency"] = unique_currencies[0] if unique_currencies else ""

    return merged


def _merge_lane_group_greedy(
    group: pd.DataFrame,
    transport_cost_columns: list[str],
) -> list[dict[str, object]]:
    """Merge rows within a lane group when their transport costs do not overlap."""
    remaining = [row.to_dict() for _, row in group.iterrows()]
    merged_rows: list[dict[str, object]] = []

    while remaining:
        base = remaining.pop(0)
        index = 0
        while index < len(remaining):
            if _can_merge_rows(base, remaining[index], transport_cost_columns):
                base = _merge_two_rows(base, remaining[index], transport_cost_columns)
                remaining.pop(index)
            else:
                index += 1
        merged_rows.append(base)

    return merged_rows


def merge_unique_lanes(matrix_df: pd.DataFrame) -> pd.DataFrame:
    """Merge rows with identical shipment info when transport costs do not conflict."""
    if matrix_df.empty:
        return matrix_df

    transport_cost_columns = transport_cost_columns_from_df(matrix_df)
    grouped_rows: dict[tuple[str, ...], list[int]] = {}

    for index, row in matrix_df.iterrows():
        key = _lane_identity_key(row)
        grouped_rows.setdefault(key, []).append(index)

    merged_rows: list[dict[str, object]] = []
    for indices in grouped_rows.values():
        group = matrix_df.loc[indices]
        if len(group) == 1:
            merged_rows.append(group.iloc[0].to_dict())
            continue
        merged_rows.extend(_merge_lane_group_greedy(group, transport_cost_columns))

    result = pd.DataFrame(merged_rows)
    result["Lane #"] = range(1, len(result) + 1)

    ordered_columns = [*SHIPMENT_COLUMNS, "Currency", *transport_cost_columns]
    return result.loc[:, ordered_columns]


def _style_header_cell(
    cell,
    *,
    bold: bool = True,
    center: bool = False,
    fill: PatternFill = HEADER_FILL,
) -> None:
    cell.font = BOLD if bold else NORMAL
    cell.fill = fill
    cell.alignment = CENTER if center else LEFT
    cell.border = THIN_BORDER


def _style_cost_meta_cell(cell, *, fill: PatternFill = COST_META_FILL) -> None:
    cell.font = NORMAL
    cell.fill = fill
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def _write_merged_header_cell(
    worksheet,
    row_index: int,
    start_col: int,
    end_col: int,
    value: str,
    *,
    fill: PatternFill,
    bold: bool = True,
    center: bool = True,
) -> None:
    if end_col > start_col:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = worksheet.cell(row_index, start_col, value)
    _style_header_cell(cell, bold=bold, center=center, fill=fill)


def _write_merged_meta_cell(
    worksheet,
    row_index: int,
    start_col: int,
    end_col: int,
    value: str,
    *,
    fill: PatternFill = COST_META_FILL,
) -> None:
    if end_col > start_col:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = worksheet.cell(row_index, start_col, value)
    _style_cost_meta_cell(cell, fill=fill)


def _column_width_for_header(header: str) -> float:
    return min(42.0, max(14.0, len(header) * 0.9 + 4))


def write_matrix_sheet(
    workbook: Workbook,
    matrix_df: pd.DataFrame,
    *,
    sheet_name: str = "Rate card",
    duplicate_row_indices: set[int] | None = None,
    service_fee_rate_by: dict[str, str] | None = None,
    service_fee_currencies: dict[str, str] | None = None,
) -> None:
    worksheet = workbook.active
    worksheet.title = sheet_name

    transport_cost_columns = transport_cost_columns_from_df(matrix_df)
    service_fee_columns = service_fee_cost_columns_from_df(matrix_df)
    service_fee_rate_by = service_fee_rate_by or {}
    service_fee_currencies = service_fee_currencies or {}
    duplicate_row_indices = duplicate_row_indices or set()

    shipment_count = len(SHIPMENT_COLUMNS)
    transport_currency_col = shipment_count + 1
    transport_rate_start_col = transport_currency_col + 1
    if transport_cost_columns:
        transport_rate_end_col = transport_rate_start_col + len(transport_cost_columns) - 1
    else:
        transport_rate_end_col = transport_currency_col

    service_fee_column_pairs: list[tuple[str, int, int]] = []
    next_col = transport_rate_end_col + 1
    for cost_name in service_fee_columns:
        service_fee_column_pairs.append((cost_name, next_col, next_col + 1))
        next_col += 2

    if service_fee_column_pairs:
        service_fee_end_col = service_fee_column_pairs[-1][2]
    else:
        service_fee_end_col = transport_rate_end_col

    if transport_cost_columns:
        _write_merged_header_cell(
            worksheet,
            COST_GROUP_ROW,
            transport_currency_col,
            transport_rate_end_col,
            "Grouped cost: Transport cost",
            fill=TRANSPORT_GROUP_FILL,
        )

        for offset, cost_name in enumerate(transport_cost_columns, start=transport_rate_start_col):
            cell = worksheet.cell(COST_NAME_ROW, offset, cost_name)
            _style_header_cell(cell, fill=TRANSPORT_COST_FILL, center=True)

        _write_merged_meta_cell(
            worksheet,
            APPLY_IF_ROW,
            transport_currency_col,
            transport_rate_end_col,
            "Applies if invoiced by Carrier",
        )

        for offset, cost_name in enumerate(transport_cost_columns, start=transport_rate_start_col):
            cell = worksheet.cell(RATE_BY_ROW, offset, rate_by_for_cost_column(cost_name))
            _style_cost_meta_cell(cell)

        currency_rate_by = worksheet.cell(RATE_BY_ROW, transport_currency_col)
        _style_cost_meta_cell(currency_rate_by)

    for cost_name, currency_col, rate_col in service_fee_column_pairs:
        _write_merged_header_cell(
            worksheet,
            COST_NAME_ROW,
            currency_col,
            rate_col,
            cost_name,
            fill=SERVICE_FEE_COST_FILL,
            center=True,
        )

        _write_merged_meta_cell(
            worksheet,
            APPLY_IF_ROW,
            currency_col,
            rate_col,
            "Applies if invoiced by Carrier",
            fill=SERVICE_FEE_META_FILL,
        )

        _write_merged_meta_cell(
            worksheet,
            RATE_BY_ROW,
            currency_col,
            rate_col,
            service_fee_rate_by.get(cost_name, "Rate by:\nRegular rule"),
            fill=SERVICE_FEE_META_FILL,
        )

    for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
        cell = worksheet.cell(COLUMN_HEADER_ROW, col_index, header)
        _style_header_cell(cell, bold=header in {"Lane #", "Service Type", "Destination City ", "Equipment Type"})

    if transport_cost_columns:
        transport_currency_header = worksheet.cell(COLUMN_HEADER_ROW, transport_currency_col, "Currency")
        _style_header_cell(transport_currency_header, center=True, fill=TRANSPORT_COST_FILL)

        for offset in range(transport_rate_start_col, transport_rate_end_col + 1):
            header_cell = worksheet.cell(COLUMN_HEADER_ROW, offset, "p/unit")
            _style_header_cell(header_cell, center=True, fill=TRANSPORT_COST_FILL)

    for cost_name, currency_col, rate_col in service_fee_column_pairs:
        currency_header = worksheet.cell(COLUMN_HEADER_ROW, currency_col, "Currency")
        _style_header_cell(currency_header, center=True, fill=SERVICE_FEE_COST_FILL)

        rate_header = worksheet.cell(COLUMN_HEADER_ROW, rate_col, "p/unit")
        _style_header_cell(rate_header, center=True, fill=SERVICE_FEE_COST_FILL)

    for matrix_index, (_, row) in enumerate(matrix_df.iterrows()):
        excel_row = DATA_START_ROW + matrix_index
        is_duplicate = matrix_index in duplicate_row_indices
        is_special = cell_text(row.get("Service Type")) == SPECIAL_SERVICE_TYPE
        row_fill = DUPLICATE_LANE_FILL if is_duplicate else None

        for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
            cell = worksheet.cell(excel_row, col_index, row.get(header))
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            if row_fill is not None:
                cell.fill = row_fill
            elif is_special:
                cell.fill = SPECIAL_ROW_FILL

        if transport_cost_columns:
            transport_currency_value = "" if is_special else row.get("Currency")
            currency_cell = worksheet.cell(excel_row, transport_currency_col, transport_currency_value)
            currency_cell.alignment = CENTER
            currency_cell.border = THIN_BORDER
            if row_fill is not None:
                currency_cell.fill = row_fill
            elif is_special:
                currency_cell.fill = SPECIAL_ROW_FILL

            for offset, cost_name in enumerate(transport_cost_columns, start=transport_rate_start_col):
                value = row.get(cost_name)
                cell = worksheet.cell(excel_row, offset)
                cell.border = THIN_BORDER
                if value is not None and value != "":
                    cell.value = value
                    cell.number_format = RATE_NUMBER_FORMAT
                    cell.alignment = CENTER
                if row_fill is not None:
                    cell.fill = row_fill
                elif is_special:
                    cell.fill = SPECIAL_ROW_FILL

        for cost_name, currency_col, rate_col in service_fee_column_pairs:
            currency_value = service_fee_currencies.get(cost_name, "") if is_special else ""
            currency_cell = worksheet.cell(excel_row, currency_col, currency_value)
            currency_cell.alignment = CENTER
            currency_cell.border = THIN_BORDER
            if row_fill is not None:
                currency_cell.fill = row_fill
            elif is_special:
                currency_cell.fill = SPECIAL_ROW_FILL

            value = row.get(cost_name)
            rate_cell = worksheet.cell(excel_row, rate_col)
            rate_cell.border = THIN_BORDER
            if value is not None and value != "":
                rate_cell.value = value
                rate_cell.number_format = RATE_NUMBER_FORMAT
                rate_cell.alignment = CENTER
            if row_fill is not None:
                rate_cell.fill = row_fill
            elif is_special:
                rate_cell.fill = SPECIAL_ROW_FILL

    for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = _column_width_for_header(header)

    if transport_cost_columns:
        worksheet.column_dimensions[get_column_letter(transport_currency_col)].width = 12.0
        for offset, cost_name in enumerate(transport_cost_columns, start=transport_rate_start_col):
            worksheet.column_dimensions[get_column_letter(offset)].width = _column_width_for_header(cost_name)

    for cost_name, currency_col, rate_col in service_fee_column_pairs:
        worksheet.column_dimensions[get_column_letter(currency_col)].width = 12.0
        worksheet.column_dimensions[get_column_letter(rate_col)].width = _column_width_for_header(cost_name)

    worksheet.freeze_panes = worksheet.cell(DATA_START_ROW, 1)
    worksheet.sheet_view.showGridLines = False


def save_matrix(
    matrix_df: pd.DataFrame,
    source_file: Path,
    *,
    output_path: Path | None = None,
    duplicate_row_indices: set[int] | None = None,
    service_fee_rate_by: dict[str, str] | None = None,
    service_fee_currencies: dict[str, str] | None = None,
    accessorial_costs: list[tuple[str, float | int, str, str, str]] | None = None,
) -> tuple[Path, int, int, int]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        output_path = OUTPUT_DIR / f"{source_file.stem.replace('_extracted', '')}_matrix.xlsx"

    workbook = Workbook()
    write_matrix_sheet(
        workbook,
        matrix_df,
        duplicate_row_indices=duplicate_row_indices,
        service_fee_rate_by=service_fee_rate_by,
        service_fee_currencies=service_fee_currencies,
    )
    if accessorial_costs:
        write_accessorial_costs_sheet(workbook, accessorial_costs)
    from build_conditions import write_conditions_sheet

    loading_count, entry_count, highlighted_ports = write_conditions_sheet(
        workbook,
        matrix_df,
        source_file,
    )
    workbook.save(output_path)
    return output_path, loading_count, entry_count, highlighted_ports


def run_build_matrix(
    *,
    source_file: Path | None = None,
    output_path: Path | None = None,
    auto: bool = False,
) -> Path:
    ensure_workspace_dirs()
    files = list_extracted_files()
    file_path = source_file
    if file_path is None:
        if auto and files:
            file_path = files[0]
            print(f"\nAuto mode: using extracted file {file_path.name}")
        else:
            file_path = select_extracted_file(files)

    rate_tabs = load_rate_tabs(file_path)
    if not rate_tabs:
        raise RuntimeError(
            f"No rate tabs found in {file_path.name}. "
            "Expected Base Freight Rates and/or BCL tabs."
        )

    print(f"\nBuilding matrix from {file_path.name}:")
    for tab_name, tab_df in rate_tabs:
        print(f"  - {tab_name}: {len(tab_df)} lane rows")

    matrix_df = build_shipment_and_cost_rows(rate_tabs)
    source_lane_count = count_source_lane_rows(rate_tabs)
    if source_lane_count != len(matrix_df):
        print(
            f"  Merged {source_lane_count} source lane rows into "
            f"{len(matrix_df)} unique lanes"
        )

    duplicate_indices = duplicate_lane_row_indices(matrix_df)
    if duplicate_indices:
        print(f"  Highlighting {len(duplicate_indices)} duplicate lane row(s) in yellow")

    service_fees_df = load_service_fees_tab(file_path)
    accessorial_costs = extract_accessorial_costs(service_fees_df) if service_fees_df is not None else []
    matrix_df, service_fee_rate_by, service_fee_currencies, service_fees = append_special_lane(
        matrix_df,
        service_fees_df,
    )
    if service_fee_rate_by:
        print(f"  Added SPECIAL lane with {len(service_fee_rate_by)} service fee cost(s)")
    if accessorial_costs:
        print(f"  Added {len(accessorial_costs)} accessorial cost(s) to Accessorial costs tab")

    saved_path, loading_count, entry_count, highlighted_ports = save_matrix(
        matrix_df,
        file_path,
        output_path=output_path,
        duplicate_row_indices=duplicate_indices,
        service_fee_rate_by=service_fee_rate_by,
        service_fee_currencies=service_fee_currencies,
        accessorial_costs=accessorial_costs,
    )
    print(f"\nSaved matrix ({len(matrix_df)} lanes) to: {saved_path}")

    from build_postal_code_zones import run_build_postal_code_zones

    print("\nBuilding postal code zones:")
    run_build_postal_code_zones(
        source_file=file_path,
        matrix_path=saved_path,
        matrix_df=matrix_df,
    )

    print(
        f"\nConditions tab: {loading_count} Port of Loading code(s), "
        f"{entry_count} Port of Entry code(s); "
        f"highlighted {highlighted_ports} port cell(s) in matrix"
    )
    return saved_path


def main() -> int:
    try:
        run_build_matrix()
        return 0
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
        return 130
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
